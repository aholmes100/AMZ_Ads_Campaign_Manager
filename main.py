import datetime
import requests
import json
import time
import gzip
import pandas as pd
from openpyxl import load_workbook
import shutil
from datetime import date, timedelta

CLIENT_ID = 'amzn1.application-oa2-client.43d152bd411b4a1fba916f51cd05d1ba'
PROFILE_ID = '2732942137323355'
CLIENT_SECRET = 'amzn1.oa2-cs.v1.989d42c9d3eb7d1de083e3fc8aec2e1a372bd9626fffac1eb669690560ad796f'
REFRESH_TOKEN = ('Atzr|IwEBIKQ_'
                 '-aii7YoY4EZohGnLYzSyFdNo4kvnLcfGQTNUYyYD1LFGOZxVdb_YAqpto'
                 'uuO4_ShBJqZXEsQ6lvlIV12C4qtPvAMOnRxqr9rmE69UvgyCpnH6xUrTe5'
                 'iN5jLtSducQ52AxNgpQZUaRiIoRcRBYO1gY9vOPD-zo8BTQbTOUBrd79eGC'
                 'YXRW5k8Ieyea7wQlv2eoJuZ1v8UQXVxeitNmkTWEc7pJ3ZouNKitTmhTjXBK'
                 'xok445S93F6Ui0rEEoOQOUejpA94c8hDu4tDcWeZXWgy8lWt0mRxmdgGRQkpo'
                 'MuoMImvvzBuM3DSLT_3jerIhZQiNzKQ2n80fFVldlAbhLMUs6MF8OiStq8-NcG'
                 'gHpoGSPhE_xj4hBNCtDLZn_zfXKHcUAjL62xKu7lOL6TZXH7ED16sczreSliB2xxfGEl07t_OV6s5-VTki4s8H24bY')


def get_access_token(client_id, client_secret, refresh_token):
    print("🔄 Requesting new access token...")
    token_url = "https://api.amazon.com/auth/o2/token"
    data = {
        "grant_type": "refresh_token",
        "client_id": client_id,
        "client_secret": client_secret,
        "refresh_token": refresh_token
    }
    response = requests.post(token_url, data=data)
    if response.status_code == 200:
        print("✅ Access token obtained successfully.")
        return response.json().get("access_token")
    else:
        print(f"❌ Failed to obtain access token: {response.status_code} - {response.text}")
        return None


def generate_report(start_date, end_date):
    print(f"📊 Requesting Amazon Ads report from {start_date} to {end_date}...")
    url = 'https://advertising-api.amazon.com/reporting/reports'
    headers = {
        'Content-Type': 'application/vnd.createasyncreportrequest.v3+json',
        'Amazon-Advertising-API-ClientId': CLIENT_ID,
        'Amazon-Advertising-API-Scope': PROFILE_ID,
        'Authorization': f'Bearer {access_token}'
    }

    data = {
        "name": "Ad Spend Report by SKU",
        "startDate": start_date,
        "endDate": end_date,
        "configuration": {
            "adProduct": "SPONSORED_PRODUCTS",
            "groupBy": ["advertiser"],
            "columns": ["campaignName", "advertisedAsin", "cost", "spend", "clicks", "impressions",
                        "campaignBudgetAmount", "sales7d", "unitsSoldClicks7d"],
            "reportTypeId": "spAdvertisedProduct",
            "timeUnit": "SUMMARY",
            "format": "GZIP_JSON"
        }
    }

    response = requests.post(url, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        report_id = response.json().get('reportId')
        print(f"✅ Report request successful. Report ID: {report_id}")
        return report_id
    else:
        print(f"❌ Failed to request report: {response.status_code} - {response.text}")
        return None


def check_report_status(report_id):
    print(f"⏳ Checking report status for Report ID: {report_id}...")
    url = f'https://advertising-api.amazon.com/reporting/reports/{report_id}'
    headers = {
        'Amazon-Advertising-API-ClientId': CLIENT_ID,
        'Amazon-Advertising-API-Scope': PROFILE_ID,
        'Authorization': f'Bearer {access_token}'
    }

    while True:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            status = response.json().get('status')
            if status == 'COMPLETED':
                print("✅ Report is ready for download.")
                return response.json().get('url')
            elif status in ['IN_PROGRESS', 'PENDING', 'PROCESSING']:
                print("⌛ Report is still processing... Checking again in 30 seconds.")
                time.sleep(30)
            else:
                print(f"❌ Unexpected report status: {status}")
                return None
        else:
            print(f"❌ Failed to check report status: {response.status_code} - {response.text}")
            return None


def download_report(report_url):
    print("📥 Downloading report...")
    response = requests.get(report_url, stream=True)
    if response.status_code == 200:
        with open('report.json.gz', 'wb') as file:
            file.write(response.raw.read())
        print("✅ Report downloaded successfully.")
    else:
        print(f"❌ Failed to download report: {response.status_code} - {response.text}")


def extract_and_save_report():
    print("📂 Extracting data from report...")
    with gzip.open('report.json.gz', 'rb') as f:
        json_bytes = f.read()

    json_str = json_bytes.decode('utf-8')
    report_data = json.loads(json_str)

    if isinstance(report_data, list):
        df = pd.DataFrame(report_data)
    elif isinstance(report_data, dict) and 'rows' in report_data:
        df = pd.DataFrame(report_data['rows'])
    else:
        print("❌ Unexpected JSON structure.")
        return

    ordered_columns = ["campaignName", "advertisedAsin", "cost", "spend", "clicks", "impressions",
                       "campaignBudgetAmount", "sales7d", "unitsSoldClicks7d"]
    df = df[ordered_columns]

    template_path = 'Ad_Spend_By_SKU_Template.xlsb'
    output_path = f'AMZ_Ad_Spend_Report_{date.today()}.xlsx'

    print(f"📄 Loading Excel template from: {template_path}...")
    wb = load_workbook(template_path)
    ws = wb["AMZ_Raw"]

    start_row = 2
    start_col = 2

    print("✏️ Writing data to Excel sheet: AMZ_Raw (Starting at B2)...")
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_path)
    print(f"✅ Report successfully saved as {output_path}.")


# Determine Start and End Dates Based on Today’s Date
today = date.today()
if today.day == 3:
    start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=16).strftime("%Y-%m-%d")
    end_date = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m-%d")
elif today.day == 17:
    start_date = today.replace(day=1).strftime("%Y-%m-%d")
    end_date = today.replace(day=15).strftime("%Y-%m-%d")
else:
    print("📅 Today is not the 3rd or the 18th. Exiting script.")
    exit()

print(f"📅 Running report for date range: {start_date} to {end_date}")

# Run Report Generation Process
access_token = get_access_token(CLIENT_ID, CLIENT_SECRET, REFRESH_TOKEN)
if access_token:
    report_id = generate_report(start_date, end_date)
    if report_id:
        report_url = check_report_status(report_id)
        if report_url:
            download_report(report_url)
            extract_and_save_report()
